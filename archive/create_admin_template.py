#!/usr/bin/env python3
"""Erstellt finanztracker_admin_template.xlsx — Admin-Sheet mit Users + Sessions."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

YELLOW    = "FFFF00"
DARK_BG   = "1A1A2E"
HEADER_FONT = Font(name="Calibri", bold=True, color=YELLOW)
BODY_FONT   = Font(name="Calibri", color="E0E0E0")
DARK_FILL   = PatternFill("solid", fgColor=DARK_BG)
ROW_FILL    = PatternFill("solid", fgColor="16213E")
ALT_FILL    = PatternFill("solid", fgColor="0F3460")

def thin_border():
    s = Side(style="thin", color="444466")
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(ws, cols, widths):
    ws.row_dimensions[1].height = 20
    for i, title in enumerate(cols, 1):
        c = ws.cell(1, i, title)
        c.font = HEADER_FONT; c.fill = DARK_FILL; c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

def style_row(ws, row, n, alt=False):
    fill = ALT_FILL if alt else ROW_FILL
    for c in range(1, n+1):
        cell = ws.cell(row, c)
        cell.fill = fill; cell.font = BODY_FONT; cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")

def readme_h(ws, row, text):
    c = ws.cell(row, 1, text)
    c.font = Font(name="Calibri", bold=True, size=12, color="1A56DB")
    c.fill = PatternFill("solid", fgColor="E8F0FE")
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 20

def readme_b(ws, row, text, indent=1):
    c = ws.cell(row, 1, text)
    c.font = Font(name="Calibri", size=11, color="111111")
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=indent+2)
    ws.row_dimensions[row].height = 16

def readme_blank(ws, row):
    ws.row_dimensions[row].height = 8

# ─── Sheet 1: README ──────────────────────────────────────────
ws_r = wb.active
ws_r.title = "README"
ws_r.sheet_properties.tabColor = "E5C07B"
ws_r.column_dimensions["A"].width = 76
ws_r.column_dimensions["B"].width = 20

title_cell = ws_r.cell(1, 1, "F-Tracker Admin-Sheet — Einrichtungsanleitung")
title_cell.font = Font(name="Calibri", bold=True, size=16, color=YELLOW)
title_cell.fill = DARK_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_r.row_dimensions[1].height = 30
ws_r.merge_cells("A1:B1")

readme_blank(ws_r, 2)
readme_h(ws_r, 3,  "SCHRITT 1 — Dieses Sheet zu Google Drive hochladen")
readme_b(ws_r, 4,  "→  Google Drive → '+Neu' → 'Datei hochladen' → diese Datei wählen")
readme_b(ws_r, 5,  "→  Rechtsklick → 'Öffnen mit Google Tabellen' (wird automatisch konvertiert)")
readme_blank(ws_r, 6)
readme_h(ws_r, 7,  "SCHRITT 2 — Admin Code.gs einfügen & deployen")
readme_b(ws_r, 8,  "→  Im Google Sheet: Erweiterungen → Apps Script")
readme_b(ws_r, 9,  "→  Admin Code.gs einfügen (in der F-Tracker App unter Tab 'Admin' verfügbar)")
readme_b(ws_r, 10, "→  Speichern (Disketten-Symbol)")
readme_b(ws_r, 11, "→  Bereitstellen → Neue Bereitstellung → Web-App")
readme_b(ws_r, 12, "→  Ausführen als: Ich  ·  Zugriff: Jeder (auch anonym)")
readme_b(ws_r, 13, "→  Berechtigungen bestätigen → Script-URL kopieren")
readme_blank(ws_r, 14)
readme_h(ws_r, 15, "SCHRITT 3 — Admin-Account anlegen")
readme_b(ws_r, 16, "→  Im Sheet 'Users': Zeile 2 ausfüllen (dein eigener Admin-Account)")
readme_b(ws_r, 17, "→  username: dein Benutzername (Kleinbuchstaben)")
readme_b(ws_r, 18, "→  password_hash: SHA-256 Hash deines Passworts (z.B. über https://emn178.github.io/online-tools/sha256.html)")
readme_b(ws_r, 19, "→  sheet_id + sheet_url: leer lassen (Admin braucht kein eigenes Daten-Sheet, oder manuell eintragen)")
readme_b(ws_r, 20, "→  created_at: Datum im Format 2026-01-01T00:00:00.000Z")
readme_b(ws_r, 21, "→  role: admin  ← WICHTIG")
readme_blank(ws_r, 22)
readme_h(ws_r, 23, "SCHRITT 4 — App verbinden")
readme_b(ws_r, 24, "→  F-Tracker App öffnen")
readme_b(ws_r, 25, "→  'Mit Username & Passwort anmelden' klicken")
readme_b(ws_r, 26, "→  ⚙ Admin-URL konfigurieren → Script-URL eintragen")
readme_b(ws_r, 27, "→  Mit Admin-Benutzername + Passwort anmelden")
readme_blank(ws_r, 28)
readme_h(ws_r, 29, "SCHRITT 5 — Benutzer einladen")
readme_b(ws_r, 30, "→  Im Admin-Panel: Einladungslink kopieren")
readme_b(ws_r, 31, "→  Link teilen (enthält Admin-URL bereits eingebettet)")
readme_b(ws_r, 32, "→  Neue Benutzer können sich direkt registrieren → ihr Sheet wird automatisch erstellt")
readme_blank(ws_r, 33)
readme_h(ws_r, 34, "WICHTIG — Sicherheit")
readme_b(ws_r, 35, "⚠  Dieses Sheet NUR für dich zugänglich halten (nicht öffentlich freigeben!)")
readme_b(ws_r, 36, "⚠  Apps Script läuft als dein Google-Account — alle User-Sheets werden in DEINEM Drive erstellt")
readme_b(ws_r, 37, "⚠  Passwörter werden als SHA-256-Hash gespeichert (nie im Klartext)")
readme_b(ws_r, 38, "⚠  Session-Token sind 30 Tage gültig und werden automatisch bereinigt")
readme_blank(ws_r, 39)
readme_h(ws_r, 40, "Sheet-Struktur")
readme_b(ws_r, 41, "Users:    username | password_hash | sheet_id | sheet_url | created_at | role | last_login")
readme_b(ws_r, 42, "Sessions: session_token | username | expires_at")

ws_r.freeze_panes = "A2"
ws_r.sheet_view.showGridLines = False

# ─── Sheet 2: Users ───────────────────────────────────────────
ws_u = wb.create_sheet("Users")
ws_u.sheet_properties.tabColor = "C8F53C"

set_header(ws_u,
    ["username", "password_hash", "sheet_id", "sheet_url", "created_at", "role", "last_login"],
    [18, 66, 40, 52, 24, 10, 24])
ws_u.freeze_panes = "A2"

# Example admin row (placeholder — user must fill in their own hash)
example_row = [
    "admin",
    "(SHA-256 deines Passworts hier eintragen)",
    "",
    "",
    "2026-01-01T00:00:00.000Z",
    "admin",
    ""
]
for j, val in enumerate(example_row, 1):
    ws_u.cell(2, j, val)
style_row(ws_u, 2, 7, alt=False)
ws_u.cell(2, 2).font = Font(name="Calibri", color="FF9F43", italic=True)

note = ws_u.cell(4, 1, "⚠  Zeile 2 ausfüllen: username, SHA-256 Passwort-Hash, role = 'admin'  |  sheet_id + sheet_url für Admin optional")
note.font = Font(name="Calibri", color="AA3300", italic=True, size=10)

ws_u.sheet_view.showGridLines = False

# ─── Sheet 3: Sessions ────────────────────────────────────────
ws_s = wb.create_sheet("Sessions")
ws_s.sheet_properties.tabColor = "61AFEF"

set_header(ws_s,
    ["session_token", "username", "expires_at"],
    [40, 18, 26])
ws_s.freeze_panes = "A2"

note2 = ws_s.cell(3, 1, "→  Wird automatisch von der App befüllt und bereinigt — nicht manuell bearbeiten.")
note2.font = Font(name="Calibri", color="777777", italic=True, size=10)

ws_s.sheet_view.showGridLines = False

out = "/home/user/finanztracker1/finanztracker_admin_template.xlsx"
wb.save(out)
print(f"Saved: {out}")
