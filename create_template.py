#!/usr/bin/env python3
"""Creates finanztracker_template.xlsx — README first, Einstellungen sheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── helpers ───────────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

YELLOW      = "FFFF00"
DARK_BG     = "1A1A2E"
HEADER_FONT = Font(name="Calibri", bold=True, color=YELLOW)
BODY_FONT   = Font(name="Calibri", color="E0E0E0")
DARK_FILL   = hex_fill(DARK_BG)
ROW_FILL    = hex_fill("16213E")
ALT_FILL    = hex_fill("0F3460")

def thin_border():
    s = Side(style="thin", color="444466")
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(ws, cols, col_widths=None):
    ws.row_dimensions[1].height = 20
    for i, (letter, title) in enumerate(cols, 1):
        c = ws.cell(1, i, title)
        c.font   = HEADER_FONT
        c.fill   = DARK_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        if col_widths:
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

def style_row(ws, row, n_cols, alt=False):
    fill = ALT_FILL if alt else ROW_FILL
    for c in range(1, n_cols+1):
        cell = ws.cell(row, c)
        cell.fill   = fill
        cell.font   = BODY_FONT
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")

# ─── Sheet 1: README (first!) ──────────────────────────────────────────────
ws_r = wb.active
ws_r.title = "README"
ws_r.sheet_properties.tabColor = "E5C07B"
ws_r.column_dimensions["A"].width = 74
ws_r.column_dimensions["B"].width = 20

# Title bar
title_cell = ws_r.cell(1, 1, "Finanztracker Setup — Anleitung")
title_cell.font      = Font(name="Calibri", bold=True, size=16, color=YELLOW)
title_cell.fill      = DARK_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_r.row_dimensions[1].height = 30
ws_r.merge_cells("A1:B1")

# Helpers — black body text on white background
def readme_h(row, text):
    c = ws_r.cell(row, 1, text)
    c.font      = Font(name="Calibri", bold=True, size=12, color="1A56DB")
    c.fill      = hex_fill("E8F0FE")
    c.alignment = Alignment(vertical="center", indent=1)
    ws_r.row_dimensions[row].height = 20

def readme_b(row, text, indent=0):
    c = ws_r.cell(row, 1, text)
    c.font      = Font(name="Calibri", size=11, color="111111")   # black
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=indent+2)
    ws_r.row_dimensions[row].height = 16

def readme_blank(row):
    ws_r.row_dimensions[row].height = 8

readme_blank(2)
readme_h(3,  "SCHRITT 1 — Diese Datei zu Google Drive hochladen")
readme_b(4,  "→  Google Drive öffnen → '+Neu' → 'Datei hochladen' → diese Datei wählen", 1)
readme_b(5,  "→  Rechtsklick auf die Datei → 'Öffnen mit Google Tabellen'", 1)
readme_b(6,  "→  Google konvertiert das Excel automatisch in ein Google Sheet", 1)
readme_blank(7)
readme_h(8,  "SCHRITT 2 — Apps Script bereitstellen")
readme_b(9,  "→  Im Google Sheet: Erweiterungen → Apps Script", 1)
readme_b(10, "→  Code.gs einfügen & speichern  (Code findest du in der F-Tracker App unter 'Sheet einrichten')", 1)
readme_b(11, "→  Bereitstellen → Neue Bereitstellung → Web-App", 1)
readme_b(12, "→  Ausführen als: Ich  ·  Zugriff: Jeder (auch anonym)", 1)
readme_b(13, "→  Berechtigungen bestätigen → Script-URL kopieren", 1)
readme_blank(14)
readme_h(15, "SCHRITT 3 — App verbinden")
readme_b(16, "→  F-Tracker öffnen", 1)
readme_b(17, "→  Script-URL in das URL-Feld einfügen → 'Verbinden' klicken", 1)
readme_blank(18)
readme_h(19, "WICHTIG")
readme_b(20, "⚠  Spalten-Reihenfolge NICHT verändern — die App erwartet exakt diese Struktur.", 1)
readme_b(21, "⚠  Tabellenblattnamen (Ausgaben, Einnahmen, Daueraufträge, Kategorien, Einstellungen) nicht umbenennen.", 1)
readme_b(22, "⚠  Zeile 1 (Header) nicht löschen oder verändern.", 1)
readme_blank(23)
readme_h(24, "Sheet-Struktur Übersicht")
readme_b(25, "Ausgaben:      ID | Datum | Beschreibung | Kategorie | Betrag | Notiz | Deleted", 1)
readme_b(26, "Einnahmen:     ID | Datum | Beschreibung | Kategorie | Betrag | Notiz | Deleted", 1)
readme_b(27, "Daueraufträge: ID | Was | Kategorie | Betrag | Intervall | Tag | Kommentar | Aktiv | nextDate | startDate | endDate | lastBooked", 1)
readme_b(28, "Kategorien:    ID | Name | Typ (ausgabe/einnahme) | Farbe (#HEX) | Sortierung", 1)
readme_b(29, "Einstellungen: App-Einstellungen (automatisch von der App befüllt — nicht manuell bearbeiten)", 1)
readme_blank(30)
readme_h(31, "Kategorien anpassen")
readme_b(32, "→  Im Sheet 'Kategorien' beliebige Zeilen hinzufügen/ändern.", 1)
readme_b(33, "→  Typ muss exakt 'ausgabe' oder 'einnahme' sein.", 1)
readme_b(34, "→  Farbe: Hex-Code mit # z.B. #FF6B35", 1)

ws_r.freeze_panes = "A2"

# ─── Sheet 2: Ausgaben ─────────────────────────────────────────────────────
ws_a = wb.create_sheet("Ausgaben")
ws_a.sheet_properties.tabColor = "FF6B35"
cols_a = [
    ("A","ID"), ("B","Datum"), ("C","Beschreibung"),
    ("D","Kategorie"), ("E","Betrag"), ("F","Notiz"), ("G","Deleted")
]
widths_a = [14, 13, 28, 18, 10, 22, 9]
set_header(ws_a, cols_a, widths_a)
ws_a.freeze_panes = "A2"
for i, row_data in enumerate([
    ["a001", "2026-01-01", "Beispiel Supermarkt", "Poschte", 24.5, "", ""],
    ["a002", "2026-01-02", "Beispiel Kaffee",     "Snack",    3.5, "", ""],
], start=2):
    for j, val in enumerate(row_data, 1):
        ws_a.cell(i, j, val)
    style_row(ws_a, i, 7, alt=(i%2==0))

# ─── Sheet 3: Einnahmen ────────────────────────────────────────────────────
ws_e = wb.create_sheet("Einnahmen")
ws_e.sheet_properties.tabColor = "3DBF6B"
cols_e = [
    ("A","ID"), ("B","Datum"), ("C","Beschreibung"),
    ("D","Kategorie"), ("E","Betrag"), ("F","Notiz"), ("G","Deleted")
]
widths_e = [14, 13, 28, 18, 10, 22, 9]
set_header(ws_e, cols_e, widths_e)
ws_e.freeze_panes = "A2"
for i, row_data in enumerate([
    ["e001", "2026-01-25", "Lohn Januar", "Siemens", 3127, "", ""],
], start=2):
    for j, val in enumerate(row_data, 1):
        ws_e.cell(i, j, val)
    style_row(ws_e, i, 7, alt=(i%2==0))

# ─── Sheet 4: Daueraufträge ────────────────────────────────────────────────
ws_d = wb.create_sheet("Daueraufträge")
ws_d.sheet_properties.tabColor = "61AFEF"
cols_d = [
    ("A","ID"), ("B","Was"), ("C","Kategorie"), ("D","Betrag"),
    ("E","Intervall"), ("F","Tag"), ("G","Kommentar"), ("H","Aktiv"),
    ("I","nextDate"), ("J","startDate"), ("K","endDate"), ("L","lastBooked")
]
widths_d = [14, 20, 14, 10, 13, 6, 22, 7, 14, 14, 14, 14]
set_header(ws_d, cols_d, widths_d)
ws_d.freeze_panes = "A2"
for i, row_data in enumerate([
    ["d001", "Miete", "Mieti", 1023, "monatlich", 26, "Wohnung", 1, "", "", "", ""],
    ["d002", "Handy Abo", "Handy", 37.5, "monatlich", 15, "Swisscom", 1, "", "", "", ""],
], start=2):
    for j, val in enumerate(row_data, 1):
        ws_d.cell(i, j, val)
    style_row(ws_d, i, 12, alt=(i%2==0))
ws_d.cell(15, 1, "Gültige Intervall-Werte: monatlich | wöchentlich | halbjährlich | jährlich | semestral").font = Font(color="777777", italic=True, size=10)

# ─── Sheet 5: Kategorien ───────────────────────────────────────────────────
ws_k = wb.create_sheet("Kategorien")
ws_k.sheet_properties.tabColor = "C678DD"
cols_k = [("A","ID"), ("B","Name"), ("C","Typ"), ("D","Farbe"), ("E","Sortierung")]
widths_k = [14, 20, 12, 12, 12]
set_header(ws_k, cols_k, widths_k)
ws_k.freeze_panes = "A2"

categories = [
    ("k001","Zmittag","ausgabe","#FF6B35",1),
    ("k002","Snack","ausgabe","#F7931E",2),
    ("k003","Ferien","ausgabe","#00D4AA",3),
    ("k004","Poschte","ausgabe","#4ECDC4",4),
    ("k005","Znacht","ausgabe","#FF6B6B",5),
    ("k006","Gschänk","ausgabe","#C678DD",6),
    ("k007","Chleider","ausgabe","#E06C75",7),
    ("k008","Technik","ausgabe","#61AFEF",8),
    ("k009","Mieti","ausgabe","#E5C07B",9),
    ("k010","Gsundheit","ausgabe","#56B6C2",10),
    ("k011","Internet","ausgabe","#98C379",11),
    ("k012","Handy","ausgabe","#ABB2BF",12),
    ("k013","Alkohol","ausgabe","#D19A66",13),
    ("k014","Essen in Reschti","ausgabe","#FF8FAB",14),
    ("k015","Rudern","ausgabe","#5E81F4",15),
    ("k016","Bildung","ausgabe","#A8E063",16),
    ("k017","Verlochet","ausgabe","#FD6E6A",17),
    ("k018","SBB","ausgabe","#E63946",18),
    ("k019","Möbel o.Ä.","ausgabe","#8B8FA8",19),
    ("k020","Gipfeli","ausgabe","#F6C90E",20),
    ("k021","Buch","ausgabe","#7BC8A4",21),
    ("k022","Sport","ausgabe","#FF9F43",22),
    ("k023","Freiziit","ausgabe","#54A0FF",23),
    ("k024","Diverses","ausgabe","#888888",24),
    ("k025","Siemens","einnahme","#C8F53C",25),
    ("k026","Twint","einnahme","#00C9A7",26),
    ("k027","Schenkung","einnahme","#FFD93D",27),
    ("k028","Übertrag","einnahme","#95E1D3",28),
    ("k029","Diverses","einnahme","#AAAAAA",29),
    ("k030","Kaution","einnahme","#4ECDC4",30),
]
for i, cat in enumerate(categories, start=2):
    for j, val in enumerate(cat, 1):
        ws_k.cell(i, j, val)
    style_row(ws_k, i, 5, alt=(i%2==0))
    hex_col = cat[3].lstrip("#")
    try:
        ws_k.cell(i, 4).fill = PatternFill("solid", fgColor=hex_col)
        ws_k.cell(i, 4).font = Font(color="111111", name="Calibri")
    except:
        pass

# ─── Sheet 6: Einstellungen ────────────────────────────────────────────────
ws_s = wb.create_sheet("Einstellungen")
ws_s.sheet_properties.tabColor = "888888"
ws_s.column_dimensions["A"].width = 20
ws_s.column_dimensions["B"].width = 80

# Header
h1 = ws_s.cell(1, 1, "Schlüssel")
h2 = ws_s.cell(1, 2, "Wert (JSON)")
for h in [h1, h2]:
    h.font      = HEADER_FONT
    h.fill      = DARK_FILL
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.border    = thin_border()
ws_s.row_dimensions[1].height = 20

# Prefill with empty profile
import json
default_profile = json.dumps({
    "ft_profile_v1": True,
    "userName": "",
    "theme": "",
    "lohnTag": 25,
    "sparziel": 0,
    "mSparziel": 0,
    "pinnedTabs": [],
    "homeWidgets": None,
    "notifSettings": {}
})
ws_s.cell(2, 1, "ft_profile_v1")
ws_s.cell(2, 2, default_profile)
style_row(ws_s, 2, 2)
ws_s.row_dimensions[2].height = 18

# Info rows
info_font = Font(name="Calibri", color="111111", size=10, italic=True)
ws_s.cell(4, 1, "HINWEIS:").font  = Font(name="Calibri", bold=True, size=10, color="AA3300")
ws_s.cell(5, 1, "→  Dieses Tabellenblatt wird automatisch von der F-Tracker App befüllt.").font = info_font
ws_s.cell(6, 1, "→  Bitte keine manuellen Änderungen vornehmen.").font = info_font
ws_s.cell(7, 1, "→  Gespeichert werden: Kacheln-Layout, Tab-Pins, Name, Theme, Einstellungen.").font = info_font

ws_s.freeze_panes = "A2"
ws_s.sheet_view.showGridLines = False

# ─── Global styling ────────────────────────────────────────────────────────
for ws in [ws_r, ws_a, ws_e, ws_d, ws_k]:
    ws.sheet_view.showGridLines = False

out = "/home/user/finanztracker1/finanztracker_template.xlsx"
wb.save(out)
print(f"Saved: {out}")
